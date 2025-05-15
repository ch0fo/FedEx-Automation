"""
Tool developed for David to run when doing his monthly Audit checks -> Verifies that audit files David provides all contain the necessary headers, so that his Excel macro doesn't crash.
"""

import sys
from typing import List, Dict
from tkinter import filedialog
import openpyxl.workbook
import openpyxl.worksheet
import openpyxl.worksheet.worksheet
import pandas as pd
import openpyxl
from pathlib import Path
import os
import numpy as np

def verify(args: List[str]) -> None:
    """
    Verifies the integrity of the given files by checking if all the required headers are present in all sheets of the files.\n
    Each entry shows the file verified, the sheet, and which headers are missing.
    The file/sheetname will not be shown if all headers were found (file validated successfully)
    """

    headers: List[str] = ['Awb Nbr', 'Duty Bill To Acct Nbr', 'Value Flg', 'Rod Flg', 'Employee (Last Mod)', 'Audit Result', 'Audit Fail Reason',
                          'Comment', 'Audit Date MM/DD/YY', 'Auditor Employee ID', 'Importer Nm', 'Entry Dt'] #holds all headers to check for each sheet, in each file
    
    # exports: Dict[str, Dict[int, tuple]] = dict() #Map of all results; exported to excel file at the end. Identifiable by filename.
    exports: Dict[int, tuple] = dict()
    exports_index = 0

    #Prompting files and verifying
    print("Please select files to verify")
    for file in filedialog.askopenfilenames(filetypes= [('Excel Files', '*.xls*')]): #Getting all files to process
        #reading current file
        filename = file.split('/')[-1] #getting the file name alone, without the full directory (preserves file extension)
        curr_wb: openpyxl.Workbook = openpyxl.load_workbook(filename=file, read_only=True)

        #iterating sheets, checking headers
        for sheetname in curr_wb.sheetnames:
            curr_sheet: openpyxl.worksheet.worksheet.Worksheet = curr_wb[sheetname] #getting current worksheet
            if curr_sheet.sheet_state != 'visible':
                print(f"Sheet '{sheetname}' for file '{filename}' is not visible, skipping")
                continue
            #getting headers in sheet
            curr_headers: List[str] = list()
            try:
                curr_headers = str(list(curr_sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]).removeprefix('(').removesuffix(')').replace('\'', '').split(', ')
            except:
                print(f"Unable to read headers for sheet '{sheetname}', file '{filename}' (defaulting to all headers missing)")

            #replacing new lines in headers with spaces
            curr_headers = [hdrr.replace('\\n', ' ') for hdrr in curr_headers]
            
            # print(f"Headers for '{filename}', '{sheetname}': {curr_headers}")
            compliance: List[str] = list()
            compliance.append(filename) #appending current filename
            compliance.append(sheetname) #appending current sheet name
            for header in headers: #iterating original headers, checking if header found in curr sheet's headers
                appended: bool = False
                if header in curr_headers: #checking perfect match (matching at list level)
                    compliance.append(np.nan)
                    continue #if perfect match found, move to next header

                #iterating curr headers, trying to find partial match (matching at string level)
                for curr_hdr in curr_headers:
                    if header in curr_hdr:
                        compliance.append(curr_hdr)
                        appended = True
                        break #breaking out of curr_headers loop, moving to checking next original header
                    
                if not appended: compliance.append('Missing') #finally, if no value was appended, mark as missing

            #saving results to exports
            exports[exports_index] = tuple(compliance)
            exports_index += 1

    #Creating pandas dataframe with results
    complete_headers: List[str] = ['File', 'Sheet'] #creating complete headers list to use for dataframe
    for hdr in headers:
        complete_headers.append(hdr) #appending all other previously used headers
    export_dataframe: pd.DataFrame = pd.DataFrame.from_dict(data=exports, orient='index', columns=complete_headers) #creating pandas dataframe from dict

    #Dropping perfect matches (all headers present in sheet)
    export_dataframe = export_dataframe.dropna(how='all', subset=headers, inplace=False)

    #Pasting results to excel file
    exports_path: Path = Path(os.getcwd())/'verify.xlsx' #getting path to save results to
    print(f"Exporting results to '{str(exports_path)}'")
    with pd.ExcelWriter(path= exports_path, engine='openpyxl', mode='w') as writer: #writing results out to excel file
        export_dataframe.to_excel(excel_writer=writer, sheet_name='verify', na_rep='', index=False)

    #Opening file when done pasting
    os.startfile(exports_path)

    return None

if __name__ == '__main__':
    args = sys.argv
    verify(args)