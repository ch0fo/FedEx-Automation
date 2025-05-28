import openpyxl as oxl
import tkinter as tk
from tkinter import filedialog
import datetime
import sys
import custom_classify_db
import win32com.client
import os
import pandas as pd
from typing import List, Dict, Set
import win32com
from main_automation_programs.tools import send_email, get_query
import datetime
import json
import time
import numpy as np
import copy

class Fetch:
    queryresults: pd.DataFrame
    oicsheet: object
    workbook: oxl.Workbook
    filepath: str
    sheet: object
    excelfilename: str
    def __init__(self):
        self._query = get_query(r"main_automation_programs\support-files\queries\oic_task.sql")
        self._validcoe = ['US', 'MX', 'PR']
        self._monthdict = {
            1: 'January',
            2: 'February',
            3: 'March',
            4: 'April',
            5: 'May',
            6: 'June',
            7: 'July',
            8: 'August',
            9: 'September',
            10: 'October',
            11: 'November',
            12: 'December'
        }
        self._vbapath = r"main_automation_programs\support-files\excel-macros\oic_pivotmacro\xl\vbaProject.bin"
        self.prev_sent_path = r"OIC_Tool\sent.json"
        self.date_string_format = r"%d-%b-%Y"
    
    def prompt_file(self):
        print("Select file with OIC info")
        root = tk.Tk()
        root.withdraw()
        return filedialog.askopenfile(filetypes=[("Excel files", "*.xlsx *.xlsm")]).name
    
    def clean(self):
        self.filepath = self.prompt_file()
        self.workbook = oxl.load_workbook(filename=self.filepath)
        self.sheet = self.workbook.active
        self.oicsheet = self.workbook.create_sheet(title= "OIC")
        self.oicsheet.append((cell.value for cell in list(self.sheet.rows)[0]))
        for row in self.sheet.iter_rows(min_col=1, max_col=15, min_row=2, values_only=True):
            cad = float(row[4])
            coe = str(row[12])
            if cad > 20 and coe not in self._validcoe:
                continue
            else:
                vals = []
                for value in row:
                    temp = value
                    if isinstance(value, str) and not "." in value:
                        try:
                            temp = int(value)
                        except:
                            pass
                    vals.append(temp)
                self.oicsheet.append(vals)
        self.save_changes()

    def query(self, args: List[str]):
        if len(args) != 1:
            try:
                days = int(args[1])
            except:
                print("Failed to convert args[1] to int")
                start_date = input("Please provide start date for Classify query ('D' for default): ")
                if start_date == "D":
                    start, end = self.get_defaultdates(None)
                else:
                    end_date = input("Please provide end date: ")
                    start, end = start_date, end_date
            else:
                start, end = self.get_defaultdates(days= days)
        else:
            start, end = self.get_defaultdates(None)
        qry = self._query.format(start_date= start, end_date = end)
        print("Waiting on query results")
        self.queryresults = custom_classify_db.execute_query(query= qry)
        self.pivot()

    def get_defaultdates(self, days: int | None) -> tuple[str, str]:
        today = datetime.date.today()
        if days:
            start_date = today - datetime.timedelta(days=days)
            end_date = today
            #formatting
            start = f"{start_date.day}-{self._monthdict[start_date.month][:3].upper()}-{str(start_date.year)[2:]}"
            end = f"{end_date.day}-{self._monthdict[end_date.month][:3].upper()}-{str(end_date.year)[2:]}"
        elif today.weekday() == 0:
            start_date = today - datetime.timedelta(days=2)
            end_date = today
            #formatting
            start = f"{start_date.day}-{self._monthdict[start_date.month][:3].upper()}-{str(start_date.year)[2:]}"
            end = f"{end_date.day}-{self._monthdict[end_date.month][:3].upper()}-{str(end_date.year)[2:]}"
        else:
            start = f"{today.day}-{self._monthdict[today.month][:3].upper()}-{str(today.year)[2:]}"
            end = start
        print(f"Using dates: '{start}', '{end}'")
        return start, end

    def save_changes(self):
        path = self.get_savepath()
        sheetnames = self.workbook.sheetnames
        for name in sheetnames:
            if name != "OIC":
                del self.workbook[name]
        self.workbook.save(path)
        os.startfile(path)

    def get_savepath(self) -> str:
        path = r"main_automation_programs\reports\OIC"
        tday = datetime.date.today()
        folder = "{}_{}".format(str(self._monthdict[tday.month]).lower(), tday.year)
        create_path = f"{path}\\{folder}"
        if not os.path.exists(create_path):
            os.mkdir(create_path)
            print(f"Path '{create_path}' created")
        extra_zero = ""
        if len(str(tday.day)) == 1:
            extra_zero = "0"
        self.excelfilename = "OIC_{}{}{}_{}.xlsx".format(self._monthdict[tday.month], extra_zero, tday.day, tday.year)
        path += "\\{}\\{}".format(folder, self.excelfilename)
        return path

    def clean_query(self) -> None:
        print("Cleaning query")
        for i, row in self.queryresults.iterrows():
            cad = float(row['cad_val'])
            coe = str(row['coe'])
            if cad > 20 and coe.upper() not in self._validcoe:
                self.queryresults.drop(i, inplace=True)

    def remove_prev_sent(self) -> None:
        """
        Checks the previously sent awbs so they are not sent again.
        """

        #Checking if a prev sent file exists, exiting if no file to check
        if not os.path.exists(self.prev_sent_path):
            print("No file exists to reference previously sent awbs, one will be created shortly for future reference.")
            return
        
        #If a file exists, extract awbs and remove all matches with prev sent.
        with open(self.prev_sent_path, 'r') as f:
            loaded_data: dict = json.load(f) #getting all rows from prev sent awbs
        sent_awbs: Dict[int, str] = {np.int64(key): value for key, value in loaded_data.items()} #converting the loaded data keys' into ints
        awbs_from_json: Set[int] = set(sent_awbs.keys()) #set of all awbs saved in json

        #Deleting matches with the current results
        awbs_from_results: Set[int] = set(self.queryresults['awb_nbr']) #getting awbs in results
        duplicated_awbs: Set[int] = awbs_from_results & awbs_from_json #finding intersection
        print(f"Found {len(duplicated_awbs)} previously sent awbs in results, dropping them.")
        self.queryresults = self.queryresults.drop_duplicates(subset=['awb_nbr'], inplace=False) #droppping dups from awbs
        self.queryresults = self.queryresults.set_index(keys='awb_nbr', drop= False, inplace=False) #setting awbs as index
        self.queryresults = self.queryresults.drop(list(duplicated_awbs), inplace=False) #dropping duplicates

    def add_sent_awbs(self) -> None:
        """
        Once previously sent awbs are removed from results, append all remaining awbs to the previously sent record.
        """

        #Adding existing sent awbs
        awbs_to_add: List[int] = self.queryresults['awb_nbr'].tolist()

        sent_awbs: Dict[int, str] = dict()
        if os.path.exists(self.prev_sent_path):
            with open(self.prev_sent_path, 'r') as f:
                loaded_data: dict = json.load(f) #getting all rows from prev sent awbs
            sent_awbs: Dict[int, str] = {int(key): value for key, value in loaded_data.items()} #converting the loaded data keys' into ints
        #Adding new awbs
        for awb in awbs_to_add:
            sent_awbs[awb] = datetime.datetime.today().strftime(self.date_string_format)

        #Deleting all awbs over 90 days old
        count_old_dropped: int = 0
        # sent_iterate: Dict[int, str] = copy.deepcopy(sent_awbs) #creating copy so we can modify original while iterating
        for awb, date in list(sent_awbs.items()):
            if datetime.datetime.strptime(date, self.date_string_format) < (datetime.datetime.today() - datetime.timedelta(days=90)): #converting stored date to a datetime object, then comparing if it is over 90 days old from today
                count_old_dropped += 1
                del sent_awbs[awb]
        print(f"Awbs sent over 90 days ago dropped: {count_old_dropped}")

        #Saving all remaining awbs back into .json file
        with open(self.prev_sent_path, 'w') as f:
            json.dump(obj=sent_awbs, fp=f, indent="")
        print(f"Added {len(awbs_to_add)} awbs to previously sent record.")

    def pivot(self):
        path = self.get_savepath()
        self.clean_query()
        self.remove_prev_sent()
        self.add_sent_awbs()

        #Checking if there is data left to process
        if len(self.queryresults.index) == 0:
            print("There are no awbs left to process, exiting and NOT creating a results file (empty results).")
            return

        with pd.ExcelWriter(path = path, engine='xlsxwriter') as writer:
            self.queryresults.to_excel(writer, sheet_name='OIC', na_rep='null', index=False)
            workbook = writer.book
            workbook.add_vba_project(self._vbapath)
            for sheet in workbook.sheetnames:
                if sheet != 'OIC':
                    del writer.sheets[sheet]
            sheet = writer.sheets['OIC']
            sheet.activate()
            macro_path = path.replace(".xlsx", ".xlsm")
            workbook.filename = macro_path
        
        # Opening and running macro
        xl = win32com.client.Dispatch('Excel.Application')
        wb = xl.Workbooks.Open(os.path.abspath(macro_path))
        xl.Application.Run(f'{self.excelfilename.replace(".xlsx", ".xlsm")}!Module1.Pivot')
        wb.Close(SaveChanges=1)
        del xl
        # os.startfile(macro_path)
        os.remove(path) #This is just for removing the original .xlsx document, but the .xlsm document that is acc being sent out will remain saved

        #Sending email with results
        # send_email(subject=f"OIC_TASK_0002 {datetime.datetime.today().strftime("%d-%b-%Y")}", files=[macro_path])
        
def main():

    
    args = sys.argv
    oic = Fetch()
    oic.query(args)
    # dnt.main(filepath)

if __name__ == "__main__":
    main()